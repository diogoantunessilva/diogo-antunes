import re
import argparse
from pathlib import Path
from collections import OrderedDict, Counter
from statistics import mean

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import BarChart, Reference

TARGET_COMMANDS = {
    "operational_data": "show equipment ont operational-data",
    "status_pon": "show equipment ont status pon",
    "optics": "show equipment ont optics",
}

COLUMN_RENAMES = {
    "operational_data": {
        "ont-idx": "ont-idx",
        "loss-of-signal": "loss-of-signal",
        "loss-of-ack": "loss-of-ack",
        "loss-of-gem": "loss-of-gem",
        "ont-disabled": "ont-disabled",
        "inactive": "inactive",
        "dying-gasp": "dying-gasp",
        "ont-olt-distance": "op_ont-olt-distance",
        "tod-sync": "tod-sync",
        "last-restart-reason": "last-restart-reason",
        "down-reason": "down-reason",
    },
    "status_pon": {
        "pon": "pon",
        "ont": "ont-idx",
        "sernum": "sernum",
        "admin status": "admin-status",
        "oper status": "oper-status",
        "olt-rx-sig level(dbm)": "status_olt-rx-sig-level(dbm)",
        "ont-olt distance(km)": "status_ont-olt-distance(km)",
        "desc1": "desc1",
        "desc2": "desc2",
        "hostname": "hostname",
    },
    "optics": {
        "ont-idx": "ont-idx",
        "rx-signal-level": "rx-signal-level",
        "tx-signal-level": "tx-signal-level",
        "ont-temperature": "ont-temperature",
        "ont-voltage": "ont-voltage",
        "laser-bias-curr": "laser-bias-curr",
        "olt-rx-sig-level": "optics_olt-rx-sig-level",
    },
}

NUMERIC_COLUMNS = {
    "op_ont-olt-distance",
    "status_olt-rx-sig-level(dbm)",
    "status_ont-olt-distance(km)",
    "rx-signal-level",
    "tx-signal-level",
    "ont-temperature",
    "ont-voltage",
    "laser-bias-curr",
    "optics_olt-rx-sig-level",
}

BAD_VALUES = {"unknown", "invalid", "n/a", "undefined", ""}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
ALERT_FILL = PatternFill("solid", fgColor="F4CCCC")
OK_FILL = PatternFill("solid", fgColor="D9EAD3")


import re

def extract_target_ip(text: str) -> str:
    # A expressão regular procura por "ssh" seguido de um IP
    matches = re.findall(r"ssh\s+(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})", text)
    
    if not matches:
        raise ValueError("Não foi possível localizar o IP de destino na linha do comando SSH.")
    
    return matches[0]  # Retorna o primeiro IP encontrado (como é único, só vai ter um)


def widths_from_separator(separator_line: str):
    return [len(part) for part in separator_line.split("+")]


def split_fixed_width(line: str, widths):
    values = []
    pos = 0
    for width in widths[:-1]:
        values.append(line[pos:pos + width].strip())
        pos += width + 1
    values.append(line[pos:].strip())
    return values


def extract_command_block(text: str, command: str) -> str:
    pattern = re.compile(re.escape(command), re.IGNORECASE)
    match = pattern.search(text)
    if not match:
        return ""

    remainder = text[match.end():]

    count_match = re.search(r"^\s*[\w-]+\s+count\s*:\s*\d+\s*$", remainder, flags=re.MULTILINE)
    if count_match:
        return remainder[:count_match.start()]

    prompt_match = re.search(r"^\s*typ:.*?#", remainder, flags=re.MULTILINE)
    if prompt_match:
        return remainder[:prompt_match.start()]

    return remainder


def parse_table_block(block: str):
    lines = [line.rstrip("\n") for line in block.splitlines() if line.strip()]
    if not lines:
        return [], []

    separator_idx = None
    for i, line in enumerate(lines):
        candidate = line.replace("+", "").strip()
        if "+" in line and candidate and set(candidate) <= {"-"}:
            separator_idx = i
            break

    if separator_idx is None:
        return [], []

    header_lines = [line for line in lines[:separator_idx] if "|" in line]
    if not header_lines:
        return [], []

    widths = widths_from_separator(lines[separator_idx])
    ncols = len(widths)

    normalized_header_lines = []
    for line in header_lines:
        parts = [part.strip() for part in line.split("|")]
        if len(parts) < ncols:
            parts.extend([""] * (ncols - len(parts)))
        elif len(parts) > ncols:
            parts = parts[:ncols]
        normalized_header_lines.append(parts)

    headers = []
    for col_idx in range(ncols):
        pieces = [row[col_idx] for row in normalized_header_lines if row[col_idx]]
        headers.append(" ".join(pieces).strip())

    rows = []
    for line in lines[separator_idx + 1:]:
        raw = line.strip()
        if not raw:
            continue
        if "count :" in raw:
            break
        if set(raw) <= {"-"} or set(raw) <= {"="}:
            continue

        values = split_fixed_width(line, widths)
        if len(values) != ncols:
            continue

        row = dict(zip(headers, values))
        if any(value for value in row.values()):
            rows.append(row)

    return headers, rows


def normalize_rows(rows, table_name):
    rename_map = COLUMN_RENAMES[table_name]
    normalized = []
    for row in rows:
        item = OrderedDict()
        for old_key, new_key in rename_map.items():
            item[new_key] = row.get(old_key, "").strip()
        normalized.append(item)
    return normalized


def convert_value(column_name, value):
    if value is None:
        return ""
    value = str(value).strip()
    if value == "":
        return ""
    if column_name in NUMERIC_COLUMNS:
        if value.lower() in BAD_VALUES:
            return value
        try:
            return float(value)
        except ValueError:
            return value
    return value


def is_number(value):
    return isinstance(value, (int, float))


def signal_classification(rx_value):
    if not is_number(rx_value):
        return "sem leitura"
    if rx_value <= -28:
        return "crítico"
    if rx_value <= -25:
        return "atenção"
    return "ok"


def build_problem_summary(row):
    issues = []
    oper_status = str(row.get("oper-status", "")).lower()
    inactive = str(row.get("inactive", "")).lower()
    down_reason = str(row.get("down-reason", "")).lower()
    last_restart = str(row.get("last-restart-reason", "")).lower()
    rx = row.get("rx-signal-level", "")
    olt_rx = row.get("status_olt-rx-sig-level(dbm)", "")
    temp = row.get("ont-temperature", "")

    if oper_status == "down":
        issues.append("oper down")
    if inactive == "yes":
        issues.append("inactive")
    if down_reason and down_reason not in {"n/a", "not-yet-detected"}:
        issues.append(f"down-reason={down_reason}")
    if last_restart and last_restart not in {"n/a", "unspecified-other"}:
        issues.append(f"restart={last_restart}")
    if is_number(rx) and rx <= -28:
        issues.append("rx baixo")
    if is_number(olt_rx) and olt_rx <= -28:
        issues.append("olt-rx baixo")
    if is_number(temp) and temp >= 70:
        issues.append("temperatura alta")

    return "; ".join(issues) if issues else "sem indícios"


def merge_tables(operational_rows, status_rows, optics_rows):
    merged = OrderedDict()

    for row in operational_rows:
        key = row["ont-idx"]
        merged.setdefault(key, OrderedDict())
        merged[key].update(row)

    for row in status_rows:
        key = row["ont-idx"]
        merged.setdefault(key, OrderedDict())
        merged[key].update(row)

    for row in optics_rows:
        key = row["ont-idx"]
        merged.setdefault(key, OrderedDict())
        merged[key].update(row)

    preferred_order = [
        "ont-idx", "pon", "sernum", "admin-status", "oper-status",
        "status_olt-rx-sig-level(dbm)", "status_ont-olt-distance(km)",
        "desc1", "desc2", "hostname",
        "loss-of-signal", "loss-of-ack", "loss-of-gem", "ont-disabled",
        "inactive", "dying-gasp", "op_ont-olt-distance", "tod-sync",
        "last-restart-reason", "down-reason",
        "rx-signal-level", "tx-signal-level", "ont-temperature",
        "ont-voltage", "laser-bias-curr", "optics_olt-rx-sig-level"
    ]

    all_columns = []
    for column in preferred_order:
        if any(column in row for row in merged.values()):
            all_columns.append(column)

    for row in merged.values():
        for column in row.keys():
            if column not in all_columns:
                all_columns.append(column)

    for col in ["classificacao_rx", "problemas_detectados"]:
        if col not in all_columns:
            all_columns.append(col)

    output_rows = []
    for _, row in merged.items():
        ordered_row = OrderedDict()
        for column in all_columns:
            ordered_row[column] = convert_value(column, row.get(column, ""))
        ordered_row["classificacao_rx"] = signal_classification(ordered_row.get("rx-signal-level", ""))
        ordered_row["problemas_detectados"] = build_problem_summary(ordered_row)
        output_rows.append(ordered_row)

    return all_columns, output_rows


def autosize_worksheet(ws):
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in column_cells:
            cell.border = THIN_BORDER
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_length + 2, 12), 40)


def style_header(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_table_sheet(wb, sheet_name, rows):
    ws = wb.create_sheet(title=sheet_name)
    if not rows:
        ws.append(["Sem dados"])
        return ws

    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    style_header(ws)
    autosize_worksheet(ws)
    return ws


def add_conditional_formatting(ws, headers):
    header_map = {name: idx + 1 for idx, name in enumerate(headers)}

    if "rx-signal-level" in header_map:
        col = get_column_letter(header_map["rx-signal-level"])
        data_range = f"{col}2:{col}{ws.max_row}"
        ws.conditional_formatting.add(data_range, CellIsRule(operator="lessThanOrEqual", formula=["-28"], fill=ALERT_FILL))
        ws.conditional_formatting.add(data_range, CellIsRule(operator="between", formula=["-28", "-25"], fill=WARN_FILL))

    if "status_olt-rx-sig-level(dbm)" in header_map:
        col = get_column_letter(header_map["status_olt-rx-sig-level(dbm)"])
        data_range = f"{col}2:{col}{ws.max_row}"
        ws.conditional_formatting.add(data_range, CellIsRule(operator="lessThanOrEqual", formula=["-28"], fill=ALERT_FILL))

    if "oper-status" in header_map:
        col = get_column_letter(header_map["oper-status"])
        rng = f"{col}2:{col}{ws.max_row}"
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'LOWER({col}2)="down"'], fill=ALERT_FILL))

    if "classificacao_rx" in header_map:
        col = get_column_letter(header_map["classificacao_rx"])
        rng = f"{col}2:{col}{ws.max_row}"
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'LOWER({col}2)="crítico"'], fill=ALERT_FILL))
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'LOWER({col}2)="atenção"'], fill=WARN_FILL))
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'LOWER({col}2)="ok"'], fill=OK_FILL))


def create_summary_sheet(wb, ip, consolidated_rows):
    ws = wb.create_sheet(title="Resumo")
    ws.append(["Indicador", "Valor"])

    total = len(consolidated_rows)
    up_count = sum(1 for r in consolidated_rows if str(r.get("oper-status", "")).lower() == "up")
    down_count = sum(1 for r in consolidated_rows if str(r.get("oper-status", "")).lower() == "down")
    inactive_count = sum(1 for r in consolidated_rows if str(r.get("inactive", "")).lower() == "yes")

    rx_values = [r.get("rx-signal-level") for r in consolidated_rows if is_number(r.get("rx-signal-level"))]
    critical_rx = sum(1 for r in consolidated_rows if r.get("classificacao_rx") == "crítico")
    warn_rx = sum(1 for r in consolidated_rows if r.get("classificacao_rx") == "atenção")

    down_reason_counter = Counter(
        str(r.get("down-reason", "")).strip()
        for r in consolidated_rows
        if str(r.get("down-reason", "")).strip() not in {"", "n/a", "not-yet-detected"}
    )

    problem_counter = Counter(
        str(r.get("problemas_detectados", "")).strip()
        for r in consolidated_rows
        if str(r.get("problemas_detectados", "")).strip() not in {"", "sem indícios"}
    )

    metrics = [
        ("IP da OLT", ip),
        ("Total de ONTs", total),
        ("ONTs UP", up_count),
        ("ONTs DOWN", down_count),
        ("ONTs inativas", inactive_count),
        ("RX crítico (<= -28 dBm)", critical_rx),
        ("RX atenção (-28 a -25 dBm)", warn_rx),
        ("Média RX sinal", round(mean(rx_values), 3) if rx_values else "n/d"),
        ("Menor RX sinal", round(min(rx_values), 3) if rx_values else "n/d"),
        ("Maior RX sinal", round(max(rx_values), 3) if rx_values else "n/d"),
    ]

    for row in metrics:
        ws.append(list(row))

    ws.append([])
    ws.append(["Top down-reasons", "Quantidade"])
    start_down_reason = ws.max_row + 1
    for reason, qty in down_reason_counter.most_common(10):
        ws.append([reason, qty])

    if not down_reason_counter:
        ws.append(["sem ocorrências", 0])

    ws.append([])
    ws.append(["Top problemas detectados", "Quantidade"])
    start_problem = ws.max_row + 1
    for problem, qty in problem_counter.most_common(10):
        ws.append([problem, qty])

    if not problem_counter:
        ws.append(["sem ocorrências", 0])

    style_header(ws)
    autosize_worksheet(ws)

    chart1 = BarChart()
    chart1.title = "Top Down Reasons"
    chart1.y_axis.title = "Quantidade"
    chart1.x_axis.title = "Motivo"

    down_end = start_down_reason + max(len(down_reason_counter), 1) - 1
    data = Reference(ws, min_col=2, min_row=start_down_reason, max_row=down_end)
    cats = Reference(ws, min_col=1, min_row=start_down_reason, max_row=down_end)
    chart1.add_data(data, titles_from_data=False)
    chart1.set_categories(cats)
    chart1.height = 8
    chart1.width = 16
    ws.add_chart(chart1, "D2")

    chart2 = BarChart()
    chart2.title = "Top Problemas Detectados"
    chart2.y_axis.title = "Quantidade"
    chart2.x_axis.title = "Problema"

    problem_end = start_problem + max(len(problem_counter), 1) - 1
    data2 = Reference(ws, min_col=2, min_row=start_problem, max_row=problem_end)
    cats2 = Reference(ws, min_col=1, min_row=start_problem, max_row=problem_end)
    chart2.add_data(data2, titles_from_data=False)
    chart2.set_categories(cats2)
    chart2.height = 8
    chart2.width = 16
    ws.add_chart(chart2, "D20")

    return ws


def create_consolidated_sheet(wb, rows):
    ws = wb.create_sheet(title="Consolidado")
    if not rows:
        ws.append(["Sem dados consolidados"])
        return ws

    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    style_header(ws)
    autosize_worksheet(ws)
    add_conditional_formatting(ws, headers)
    return ws


def parse_log_file(log_path: Path):
    text = log_path.read_text(encoding="utf-8", errors="ignore")
    print(text[:500])  # Exibe os primeiros 500 caracteres para inspeção
    ip = extract_target_ip(text)
    

    parsed_tables = {}
    for table_name, command in TARGET_COMMANDS.items():
        block = extract_command_block(text, command)
        _, rows = parse_table_block(block)
        parsed_tables[table_name] = normalize_rows(rows, table_name)

    headers, consolidated_rows = merge_tables(
        parsed_tables["operational_data"],
        parsed_tables["status_pon"],
        parsed_tables["optics"],
    )

    return {
        "ip": ip,
        "headers": headers,
        "consolidated": consolidated_rows,
        "operational_data": parsed_tables["operational_data"],
        "status_pon": parsed_tables["status_pon"],
        "optics": parsed_tables["optics"],
    }


def generate_workbook(parsed_data: dict, output_path: Path):
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    create_consolidated_sheet(wb, parsed_data["consolidated"])
    write_table_sheet(wb, "operational_data", parsed_data["operational_data"])
    write_table_sheet(wb, "status_pon", parsed_data["status_pon"])
    write_table_sheet(wb, "optics", parsed_data["optics"])
    create_summary_sheet(wb, parsed_data["ip"], parsed_data["consolidated"])

    wb.save(output_path)


def process_one_file(log_path: Path, output_dir: Path):
    parsed = parse_log_file(log_path)
    
    # Extrai o nome do comando de acordo com o que foi encontrado no arquivo
    # Isso irá pegar o nome do comando (por exemplo, "operational_data", "status_pon", "optics")
    command_name = [cmd for cmd, command in TARGET_COMMANDS.items() if command in parsed["ip"]][0]
    
    # Usa o comando e o IP para formar o nome do arquivo
    output_path = output_dir / f"{command_name}_{parsed['ip']}.xlsx"
    generate_workbook(parsed, output_path)
    return output_path

def process_directory(input_dir: Path, output_dir: Path):
    generated = []
    for log_path in sorted(input_dir.glob("*.txt")):
        try:
            generated.append(process_one_file(log_path, output_dir))
        except Exception as exc:
            print(f"[ERRO] {log_path.name}: {exc}")
    return generated


def build_arg_parser():
    parser = argparse.ArgumentParser(
        description="Parser avançado de logs Nokia OLT para Excel por IP da sessão SSH."
    )
    parser.add_argument(
        "input_path",
        help="Arquivo .txt de log ou pasta contendo vários .txt"
    )
    parser.add_argument(
        "-o", "--output-dir",
        default=".",
        help="Pasta de saída dos arquivos .xlsx"
    )
    return parser


def main():
    # Caminho fixo para o arquivo ou diretório de entrada
    input_path = Path(r"C:\teste\teste.log")
    output_dir = Path(r"C:\teste")

    output_dir.mkdir(parents=True, exist_ok=True)

    if input_path.is_file():
        generated = [process_one_file(input_path, output_dir)]
    elif input_path.is_dir():
        generated = process_directory(input_path, output_dir)
    else:
        raise FileNotFoundError(f"Caminho não encontrado: {input_path}")

    if not generated:
        print("Nenhuma planilha foi gerada.")
        return

    print("Arquivos gerados:")
    for path in generated:
        print(f" - {path}")


if __name__ == "__main__":
    main()